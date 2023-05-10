"""Defined an automator model"""
import os
import pandas as pd
from datetime import timedelta
from fredapi import Fred
from openpyxl import Workbook, load_workbook


class FredData:
    def __init__(self, api_key):
        self.api_key = api_key
        self.fred = Fred(api_key='[my_api_key]')
        self.latest_data = None
        self.last_year_data = None

    def get_latest_data(self, series_names):
        """
        Get latest data from FRED API for given series names
        """
        latest_data = []
        for series_name in series_names:
            try:
                data = self.fred.get_series(series_name)
                index_counter = -1
                # check if data is not nan
                while not isinstance(data[index_counter], (int, float)):
                    index_counter -= 1
                # assign the last (:latest) row date (index equals date: .index[n] returns date on line n)
                latest_date = data.index[index_counter]
                # assign the last data. (data[n]: returns data on line n)
                latest_value = data[index_counter]
                latest_data.append((series_name, latest_date, latest_value))
            except Exception as e:
                print(f"Failed to get latest data for {series_name}: {e}")
        self.latest_data = pd.DataFrame(latest_data, columns=['Series', 'Date', 'Value'])

    def get_last_year_data(self, series_names):
        """
        Get last year data from FRED API for given series names on the same date as latest data
        """
        last_year_data = []
        for series_name in series_names:
            try:
                data = self.fred.get_series(series_name)
                """
                call five series data from get_latest_data and return a Numpy representation which is 'Date' value.
                .values[0]: Only the values in the DataFrame will be returned, the axes labels will be removed.
                .loc[] returns  'Date' column series
                """
                latest_date = self.latest_data.loc[self.latest_data['Series'] == series_name, 'Date'].values[0]
                last_year_date = pd.to_datetime(latest_date).to_pydatetime() - timedelta(days=365)
                # plus 1 day until the valid last year date is found in index and check if data is not nan
                while last_year_date not in data.index:
                    last_year_date += timedelta(days=1)
                    if isinstance([last_year_date], (int, float)):
                        break
                last_year_value = data[last_year_date]
                last_year_data.append((series_name, last_year_date, last_year_value))
            except Exception as e:
                print(f"Failed to get last year data for {series_name}: {e}")
        self.last_year_data = pd.DataFrame(last_year_data, columns=['Series', 'Date', 'Value'])

    def write_to_excel(self, filepath, cell_locations):
        """
        Write latest and last year data to Excel file
        """
        if not os.path.exists(filepath):
            wb = Workbook()
            wb.save(filepath)
        wb = load_workbook(filepath, data_only=True)
        ws = wb['terminal']
        
        """Write latest data: date and value"""
        for series_name in cell_locations:
            cell_location = cell_locations[series_name]['latest']
            date_from_latest_date = self.latest_data.loc[self.latest_data['Series'] == series_name, 'Date'].values[0]
            ws[cell_location] = pd.to_datetime(date_from_latest_date).to_pydatetime().strftime('%Y/%m/%d')
            ws[chr(ord(cell_location[0]) + 1) + str(int(cell_location[1:]))] = \
                self.latest_data.loc[self.latest_data['Series'] == series_name, 'Value'].values[0]

        # Write last year data
        for series_name in cell_locations:
            cell_location = cell_locations[series_name]['last_year']
            ws[cell_location] = self.last_year_data.loc[self.last_year_data['Series'] == series_name, 'Value'].values[0]

        wb.save(filepath)
        print(f"Data written to {filepath}.")
